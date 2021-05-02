import openpyxl
import os
import re

from Stock import Stock
from datetime import datetime
from Config import Config

class TableWorker:
    def __new__(cls, config, *args, **kwargs):
        if not hasattr(cls, 'instance') and config != None:
            cls.instance = super(TableWorker, cls).__new__(cls, *args, **kwargs)
        return cls.instance
    
    def __init__(self, config, *args, **kwargs):
        self.Path = config.TablePath
        self.Config = config
        self.Days = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота", "воскресенье"]

        while os.path.exists(self.Path) is False:
            self.Path = input("Введите путь до таблицы: ").replace('\"', '')
            config.TablePath = self.Path
        
        self.Book = openpyxl.load_workbook(self.Path)
        self.Sheet = self.Book.active
        self.EndTickerPos = list()
        TableData = self.GetTickersAndDate()
        self.Tickers = TableData[0]
        self.RealDate = TableData[1]
        self.Date = f"{self.RealDate.month}/{self.RealDate.day}/{self.RealDate.year}"
        print(f"Поиск за {self.RealDate.day}/{self.RealDate.month}/{self.RealDate.year}, это - {self.Days[self.RealDate.weekday()]}")
        print(f"акции - {', '.join([tic.Ticker for tic in self.Tickers])}")
    
    def GetTickersAndDate(self):
        Tickers = list()
        Date = datetime(2021, 3, 22)    
        CashTickers = self.Config.StocksId_cash.keys()


        for position in self.Config.TickersPoses:
            ticker = self.Sheet.cell(row=position[0], column=position[1]).value

            if ticker != None and TableWorker.IsTicker(ticker):
                    if TableWorker.IsPrivilegedTicker(ticker):
                        ticker = TableWorker.MakePrivilegedTicker(ticker)
                    
                    country = self.Sheet.cell(row=position[0], column=position[1] + 1).value
                    id_ = None
                    
                    if ticker in CashTickers:
                        id_ = self.Config.StocksId_cash[ticker]
                    
                    Tickers.append(Stock(ticker, country, id_))
    
        return (Tickers, Date)
    
    def WriteTickersPrice(self, data:dict):
        for position in self.Config.TickersPoses:
            cell = self.Sheet.cell(row=position[0], column=position[1])

            if cell is not None and cell.value is not None:
                ticker = cell.value

                if TableWorker.IsPrivilegedTicker(ticker):
                        ticker = TableWorker.MakePrivilegedTicker(ticker)

                if ticker in data.keys():
                    self.Sheet.cell(row=position[0], column=position[1] + 2).value = data[ticker]

        path = self.Config.TablePath
        if '.xlsm' in path:
            path = path[:path.index('.xlsm')] + '_buffer.xlsx'

        self.Book.save(path)
        os.startfile(path)
        print("Таблица успешно изменена")

    @staticmethod
    def IsPrivilegedTicker(ticker:str):
        if '_p' in ticker or '_P' in ticker:
            return True
        
        return False

    @staticmethod
    def MakePrivilegedTicker(ticker:str):
        if '_P' in ticker:
            ticker = ticker[:ticker.index('_P')] + '_p'

        return ticker

    @staticmethod
    def IsСoordinates(value:list):
        FirstCond = len(value) > 1

        if FirstCond is False:
            return False

        SecondCond = value[0] > 0 and value[1] > 0
        ThirdCond = value[0] < 1048576 and value[1] < 1048576

        return SecondCond and ThirdCond

    @staticmethod
    def IsTicker(value:str):
        return re.search(r'[^A-Z0-9_p]', value) == None