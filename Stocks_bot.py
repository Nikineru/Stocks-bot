import requests
import fake_useragent
import openpyxl
import re
import datetime
import calendar
import ID_founder
from time import time
from bs4 import BeautifulSoup

#ID_founder.main(file_name = r"C:\Users\isbud\OneDrive\Рабочий стол\Stocks.txt", text = 'russia GAZP')
user = fake_useragent.UserAgent().random
session = requests.Session()


def GetStockData(Stock_id, search_date):
    start = time()

    print(f"ID - {Stock_id}") 
    params = {
            "curr_id": Stock_id,
            "smlID": 0,
            "header": f'Прошлые данные - {ticker}',
            'st_date': search_date,
            'end_date': search_date,
            "interval_sec": 'Daily',
            "sort_col": "date",
            "sort_ord": "DESC",
            "action": "historical_data"
        }

    head = {
        "User-Agent": user,
        "X-Requested-With": "XMLHttpRequest",
        "Accept": "text/html",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    }

    URL = "https://www.investing.com/instruments/HistoricalDataAjax"
    Period_Data = requests.post(URL, headers=head, data=params)

    soup = BeautifulSoup(Period_Data.content, 'html.parser')
    StocksQuote = soup.find("td", { "class" : re.compile(r"^(greenFont|redFont)$") })['data-real-value']

    print("Время - {:.2F}".format(time() - start))
    return StocksQuote

def GetStockCountry(Ticker):
    URL = f'https://ru.investing.com/search/?q={Ticker}'

    header = {
        'user-agent': user
    }

    response = requests.get(URL, headers=header)
    soup = BeautifulSoup(response.content, 'html.parser')

    FirstStock = soup.find('a', class_ = 'js-inner-all-results-quote-item row')
    if FirstStock != None:
        URL = 'https://www.investing.com{}'.format(FirstStock['href'])
    else:
        print("I can`t frind stock data")

    response = requests.get(URL, headers=header)
    soup = BeautifulSoup(response.content, 'html.parser')
    country_data = soup.find_all('a', class_='font-bold text-secondary hover:underline')
    return country_data[2].text.lower()

def GetAmericanDate(date):
    return f"{GetStringFromDay(date.month)}/{GetStringFromDay(date.day)}/{GetStringFromDay(date.year)}"

def GetStringFromDay(value):
    return value if value >= 10 else f"0{value}"

def IsTicker(value:str):
    value = value.replace(' ', '')

    if re.search(r'[^A-Z]', value):
        return False
    else:
        return True

def GetInputPosition(sheet, path=r"C:\Users\isbud\OneDrive\Рабочий стол\Stocks-bot\InputData.txt"):
    DatePos = None
    TickerPos = None
    
    File = open(path, 'r')

    lines = File.readlines()
    if len(lines) > 0:
        _date = [int(i) for i in lines[0].split()]
        _ticker = [int(i) for i in lines[1].split()]
        
        if type(sheet.cell(row=_date[0], column=_date[1]).value) == datetime.datetime:
            DatePos = tuple(_date)

        if IsTicker(str(sheet.cell(row=_ticker[0], column=_ticker[1]).value)):
            TickerPos = tuple(_ticker)

    File.close()

    if DatePos == None or TickerPos == None:
        for row in sheet.rows:
            for cell in row:
                if TickerPos == None and IsTicker(str(cell.value)):
                    TickerPos = (cell.row, cell.column)

                if DatePos == None and type(cell.value) == datetime.datetime:
                    DatePos = (cell.row, cell.column)
                
                if TickerPos != None and DatePos != None:
                    break
    
        File = open(path, 'w')
        date = f"{DatePos[0]} {DatePos[1]}"
        ticker = f"{TickerPos[0]} {TickerPos[1]}"

        File.write(f"{date}\n{ticker}")
        File.close()

    return (DatePos, TickerPos)

path = r"C:\Users\isbud\OneDrive\Рабочий стол\Stocks-bot\Акции.xlsx"
book = openpyxl.load_workbook(path) 
sheet = book.active

Tickers = list()
RowOfLastTicker = 1

Input = GetInputPosition(sheet)
for row in sheet.iter_rows(min_row=Input[1][0], max_col=1):
    for cell in row:
        value = str(cell.value)
        if value.isupper() and value.isdigit() == False:
            country = sheet.cell(row=cell.row, column=cell.column + 1)
            Tickers.append((cell, country))
            RowOfLastTicker += 1

Date = sheet.cell(row=Input[0][0], column=Input[0][1]).value
StartDate = GetAmericanDate(Date)
if Date > datetime.datetime.now():
    print("Введите корректную дату")
    exit()

EndDate = StartDate
print(StartDate, EndDate)
print([i[0].value for i in Tickers])

for ticker_data in Tickers:
    ticker = ticker_data[0]
    country = ticker_data[1]
    print(f'Work with {ticker.value}, {country.value}')
    
    if country.value == None:
        country.value = GetStockCountry(ticker.value)
    
    StockID = ID_founder.GetStockData(ticker=ticker.value, country=country.value)
    Quote = GetStockData(StockID, StartDate)
    sheet.cell(row=ticker.row, column=ticker.column + 2).value = Quote
book.save(path)